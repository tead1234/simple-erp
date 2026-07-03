import threading
from fastapi import Depends, HTTPException, UploadFile
from sqlalchemy.exc import IntegrityError
from app.domain.product.entity import Product
from app.domain.product.repository import IProductRepository
from app.infrastructure.database.repositories import get_product_repo, SqlProductRepository
from app.infrastructure.database.session import SessionLocal

_import_progress = {
    "running": False, "total": 0, "processed": 0,
    "created": 0, "updated": 0, "skipped": 0, "done": True, "error": None,
}


class InventoryService:
    def __init__(self, repo: IProductRepository):
        self.repo = repo

    def list(self, category=None) -> list:
        products = self.repo.list()
        if category:
            products = [p for p in products if p.category == category]
        return [self._to_dict(p) for p in products]

    def get(self, product_id: int) -> dict:
        p = self.repo.get(product_id)
        if not p:
            raise HTTPException(404, "상품을 찾을 수 없습니다")
        return self._to_dict(p)

    def create(self, name: str, code: str = None, old_code: str = None, category: str = None, model: str = None,
               stock_quantity: int = 0, min_stock_quantity: int = 0, unit_price: float = 0,
               dealer_price: float = 0, center_price: float = 0, consumer_price: float = 0) -> dict:
        try:
            p = self.repo.save(Product(
                name=name, code=code, old_code=old_code, category=category, model=model,
                stock_quantity=stock_quantity, min_stock_quantity=min_stock_quantity,
                unit_price=unit_price, dealer_price=dealer_price,
                center_price=center_price, consumer_price=consumer_price,
            ))
        except IntegrityError:
            raise HTTPException(400, "이미 등록된 품번입니다")
        return self._to_dict(p)

    def search(self, q: str, limit: int = 30) -> list:
        q = (q or "").strip()
        if not q:
            return [self._to_dict(p) for p in self.repo.recent(limit)]
        return [self._to_dict(p) for p in self.repo.search(q, limit)]

    def update(self, product_id: int, **kwargs) -> dict:
        p = self.repo.get(product_id)
        if not p:
            raise HTTPException(404, "상품을 찾을 수 없습니다")
        for k, v in kwargs.items():
            if v is not None and hasattr(p, k):
                setattr(p, k, v)
        try:
            p = self.repo.save(p)
        except IntegrityError:
            raise HTTPException(400, "이미 등록된 품번입니다")
        return self._to_dict(p)

    def adjust_stock(self, product_id: int, movement_type: str, quantity: int, reason: str = None) -> dict:
        p = self.repo.get(product_id)
        if not p:
            raise HTTPException(404, "상품을 찾을 수 없습니다")
        if movement_type == "출고" and p.stock_quantity < quantity:
            raise HTTPException(400, "재고가 부족합니다")
        self.repo.add_movement(product_id, movement_type, quantity, reason)
        if movement_type == "입고":
            p.stock_quantity += quantity
        else:
            p.stock_quantity -= quantity
        p = self.repo.save(p)
        return self._to_dict(p)

    def find_low_stock(self) -> list:
        return [self._to_dict(p) for p in self.repo.find_low_stock()]

    def start_import(self, file_bytes: bytes) -> dict:
        if _import_progress["running"]:
            raise HTTPException(409, "이미 엑셀 가져오기가 진행 중입니다")
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        if "복사 붙혀넣기" not in wb.sheetnames:
            raise HTTPException(400, "올바른 양식의 파일이 아닙니다 ('복사 붙혀넣기' 시트를 찾을 수 없습니다)")
        ws = wb["복사 붙혀넣기"]
        total = max(ws.max_row - 2, 0)  # min_row=3 기준
        if "판가 데이터" in wb.sheetnames:
            total += max(wb["판가 데이터"].max_row - 3, 0)  # min_row=4 기준
        _import_progress.update(running=True, total=total, processed=0,
                                 created=0, updated=0, skipped=0, done=False, error=None)
        threading.Thread(target=self._run_import, args=(file_bytes,), daemon=True).start()
        return {"started": True, "total": total}

    def get_import_progress(self) -> dict:
        return dict(_import_progress)

    def _run_import(self, file_bytes: bytes) -> None:
        import openpyxl, io
        db = SessionLocal()
        repo = SqlProductRepository(db)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb["복사 붙혀넣기"]
            seen_codes = set()
            processed = 0

            for row in ws.iter_rows(min_row=3, values_only=True):
                code = row[1]
                old_code = row[2]
                name = row[3]
                if code and name:
                    code = str(code).strip()
                    old_code = str(old_code).strip() if old_code else None
                    name = str(name).strip()
                    if code in seen_codes:
                        _import_progress["skipped"] += 1
                    else:
                        seen_codes.add(code)
                        model = str(row[4]).strip() if row[4] else None
                        dealer_price = float(row[6]) if row[6] else 0.0
                        center_price = float(row[7]) if row[7] else 0.0
                        consumer_price = float(row[8]) if row[8] else 0.0

                        existing = repo.find_by_code(code)
                        if existing:
                            existing.name = name
                            existing.old_code = old_code
                            existing.model = model
                            existing.dealer_price = dealer_price
                            existing.center_price = center_price
                            existing.consumer_price = consumer_price
                            repo.save(existing)
                            _import_progress["updated"] += 1
                        else:
                            repo.save(Product(
                                name=name, code=code, old_code=old_code, model=model,
                                dealer_price=dealer_price,
                                center_price=center_price, consumer_price=consumer_price,
                            ))
                            _import_progress["created"] += 1
                processed += 1
                _import_progress["processed"] = processed
                if processed % 200 == 0:
                    db.commit()

            # "판가 데이터" 시트: 구조가 달라 신품번/구품번/품명/대리점가/센터가/수요자가 6열만 존재.
            # 시트1에 없던 품번을 신규 등록하는 용도로만 쓰고, 시트1과 겹치는 품번은 시트1 값을 우선한다.
            if "판가 데이터" in wb.sheetnames:
                for row in wb["판가 데이터"].iter_rows(min_row=4, values_only=True):
                    code = row[0]
                    name = row[2]
                    if code and name:
                        code = str(code).strip()
                        name = str(name).strip()
                        if code in seen_codes:
                            _import_progress["skipped"] += 1
                        else:
                            seen_codes.add(code)
                            existing = repo.find_by_code(code)
                            if existing:
                                _import_progress["skipped"] += 1
                            else:
                                old_code = str(row[1]).strip() if row[1] else None
                                dealer_price = float(row[3]) if row[3] else 0.0
                                center_price = float(row[4]) if row[4] else 0.0
                                consumer_price = float(row[5]) if row[5] else 0.0
                                repo.save(Product(
                                    name=name, code=code, old_code=old_code,
                                    dealer_price=dealer_price,
                                    center_price=center_price, consumer_price=consumer_price,
                                ))
                                _import_progress["created"] += 1
                    processed += 1
                    _import_progress["processed"] = processed
                    if processed % 200 == 0:
                        db.commit()

            db.commit()
        except Exception as e:
            db.rollback()
            _import_progress["error"] = str(e)
        finally:
            db.close()
            _import_progress["running"] = False
            _import_progress["done"] = True

    def _to_dict(self, p: Product) -> dict:
        return {**p.__dict__, "is_low_stock": p.is_low_stock}


def get_inventory_service(repo: IProductRepository = Depends(get_product_repo)) -> InventoryService:
    return InventoryService(repo)
